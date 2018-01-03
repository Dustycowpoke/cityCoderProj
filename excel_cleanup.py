# python3 
# excel_cleanup
# removes last ; in 'location' column

import openpyxl

wb_in = 'working_stop_box.xlsx'

wb = openpyxl.load_workbook(wb_in)
sheet_in = wb.get_sheet_by_name('stopbox_final')

sheet_in['D1'] = 'loc_edit'

for row_iter in range(2, (sheet_in.max_row+1)):
	print("Row %i" % row_iter)
	loc = sheet_in.cell(row = row_iter, column = 3).value
	
	if not loc:
		pass
	
	elif (loc[-1] == ";"):
		loc_edit = loc[:-2]
	
	else:
		loc_edit = loc
		
	sheet_in.cell(row = row_iter, column = 4).value = loc_edit

wb.save('stopbox_final_2.xlsx')
print('Done')